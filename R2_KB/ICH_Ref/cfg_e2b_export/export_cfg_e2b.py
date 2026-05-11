#!/usr/bin/env python3
"""
Export CFG_E2B using exact PROFILE filter:

Outputs under --output-dir:
1) clob_exports/*.txt                 (one file per CLOB value)
2) CFG_E2B_non_clob_filtered.xlsx     (all non-CLOB columns)
3) CFG_E2B_clob_manifest_filtered.xlsx (row-to-CLOB filename mapping)
"""

from pathlib import Path
import argparse
import re

import oracledb
from openpyxl import Workbook


def sanitize(value: str, max_len: int = 100) -> str:
    raw = "" if value is None else str(value).strip()
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", raw).strip("._-")
    return (safe or "EMPTY")[:max_len]


def to_text(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "read"):  # CLOB/NCLOB
        return value.read() or ""
    return str(value)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export CFG_E2B CLOB + non-CLOB data")
    parser.add_argument("--host", required=True, help="Oracle host")
    parser.add_argument("--port", type=int, default=1521, help="Oracle port")
    parser.add_argument("--service", required=True, help="Oracle service name")
    parser.add_argument("--user", required=True, help="Oracle username")
    parser.add_argument("--password", required=True, help="Oracle password")
    parser.add_argument("--owner", default="ESM_OWNER", help="Schema owner")
    parser.add_argument("--table", default="CFG_E2B", help="Table name")
    parser.add_argument(
        "--profile",
        default="ICH-ICSR V2.1 MESSAGE TEMPLATE",
        help="Exact PROFILE filter value",
    )
    parser.add_argument(
        "--key-col",
        default="DTD_ELEMENT",
        help="Column used in CLOB filename prefix",
    )
    parser.add_argument(
        "--output-dir",
        required=True,
        help="Output folder path (e.g. .../cfg_e2b_export)",
    )
    args = parser.parse_args()

    out_dir = Path(args.output_dir)
    clob_dir = out_dir / "clob_exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    clob_dir.mkdir(parents=True, exist_ok=True)

    for fp in clob_dir.glob("*.txt"):
        fp.unlink()

    dsn = f"{args.host}:{args.port}/{args.service}"
    conn = oracledb.connect(user=args.user, password=args.password, dsn=dsn)

    meta_cur = conn.cursor()
    meta_cur.execute(
        """
        SELECT column_name, data_type, column_id
        FROM all_tab_columns
        WHERE owner = :owner AND table_name = :table
        ORDER BY column_id
        """,
        owner=args.owner.upper(),
        table=args.table.upper(),
    )
    meta = meta_cur.fetchall()
    if not meta:
        raise RuntimeError("No metadata found for table")

    all_cols = [r[0] for r in meta]
    clob_cols = [r[0] for r in meta if r[1] in ("CLOB", "NCLOB")]
    non_clob_cols = [c for c in all_cols if c not in clob_cols]
    col_idx = {c: i for i, c in enumerate(all_cols)}

    if args.key_col not in col_idx:
        raise RuntimeError(f"{args.key_col} not found in {args.owner}.{args.table}")
    if "PROFILE" not in col_idx:
        raise RuntimeError("PROFILE column not found in table")

    non_clob_path = out_dir / f"{args.table}_non_clob_filtered.xlsx"
    manifest_path = out_dir / f"{args.table}_clob_manifest_filtered.xlsx"

    wb_non = Workbook(write_only=True)
    ws_non = wb_non.create_sheet("non_clob")
    ws_non.append(non_clob_cols)

    wb_manifest = Workbook(write_only=True)
    ws_manifest = wb_manifest.create_sheet("clob_manifest")
    manifest_headers = ["row_num", args.key_col, "PROFILE"] + [f"{c}_file" for c in clob_cols]
    ws_manifest.append(manifest_headers)

    qcols = ", ".join([f'"{c}"' for c in all_cols])
    sql = f"""
        SELECT {qcols}
        FROM "{args.owner}"."{args.table}"
        WHERE "PROFILE" = :profile
        ORDER BY "DTD_ELEMENT", "DATA_ELEMENT" NULLS LAST
    """

    cur = conn.cursor()
    cur.arraysize = 1
    cur.execute(sql, profile=args.profile)

    row_num = 0
    clob_file_count = 0
    while True:
        row = cur.fetchone()
        if row is None:
            break

        row_num += 1
        key_val = to_text(row[col_idx[args.key_col]])
        profile_val = to_text(row[col_idx["PROFILE"]])
        key_safe = sanitize(key_val, 70)
        profile_safe = sanitize(profile_val, 30)

        ws_non.append([to_text(row[col_idx[c]]) for c in non_clob_cols])

        man_row = [row_num, key_val, profile_val]
        for clob_col in clob_cols:
            text = to_text(row[col_idx[clob_col]])
            file_name = ""
            if text:
                suffix = f"{profile_safe}_{sanitize(clob_col, 30)}_{row_num:05d}"
                file_path = clob_dir / f"{key_safe}_{suffix}.txt"
                bump = 1
                while file_path.exists():
                    file_path = clob_dir / f"{key_safe}_{suffix}_{bump}.txt"
                    bump += 1
                file_path.write_text(text, encoding="utf-8")
                file_name = file_path.name
                clob_file_count += 1
            man_row.append(file_name)
        ws_manifest.append(man_row)

    wb_non.save(non_clob_path)
    wb_manifest.save(manifest_path)

    print(f"rows={row_num}")
    print(f"profile_filter={args.profile}")
    print(f"clob_cols={clob_cols}")
    print(f"clob_files={clob_file_count}")
    print(f"non_clob_excel={non_clob_path}")
    print(f"manifest_excel={manifest_path}")
    print(f"clob_dir={clob_dir}")

    cur.close()
    meta_cur.close()
    conn.close()


if __name__ == "__main__":
    main()
